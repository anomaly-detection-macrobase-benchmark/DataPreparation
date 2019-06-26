import uuid

from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle, Font
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


def xlref(row, column):
    """
    Converts numerical row/column indexes (starting from 1) to Excel coordinate.
    For example, B8 (for 8, 2)
    """
    if row < 1:
        raise ValueError('Invalid row ' + row)
    if column < 1:
        raise ValueError('Invalid column ' + column)
    return get_column_letter(column) + str(row)


def xlref_range(start_row, start_column, end_row, end_column):
    """
    Converts numerical start/end row/column indexes (starting from 1) to Excel range.
    For example, B8:G16 (for 8, 2, 16, 7)
    """
    return xlref(start_row, start_column) + ':' + xlref(end_row, end_column)


def xlref_range_count(start_row, start_column, row_count, column_count):
    """
    Converts numerical start row/column indexes (starting from 1) + row/column count (like a table) to Excel range.
    For example, B8:G16 (for 8, 2, 9, 6)
    """
    if row_count < 1:
        raise ValueError('Invalid row count ' + row_count)
    if column_count < 1:
        raise ValueError('Invalid column count ' + column_count)
    return xlref_range(start_row, start_column, start_row + row_count - 1, start_column + column_count - 1)


def cell_value_to_text(value):
    if value is None:
        return ""
    return str(value)


def autofit(sheet):
    min_width = 10
    for column_cells in sheet.columns:
        col = column_cells[0].column
        length = max(len(cell_value_to_text(cell.value)) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(col)].width = max(length + 1, min_width)


def append_blank_rows(sheet, count):
    for _ in range(count):
        sheet.append([])


def append_blank_row(sheet):
    append_blank_rows(sheet, 1)


def pandas_dataframe_to_rows(df, index=True, header=True, top_left_value='') -> list:
    df_rows = list(dataframe_to_rows(df, index=index, header=header))
    if index:
        df_rows.pop(1)  # there is an empty row, but apparently it's not a bug https://groups.google.com/forum/#!topic/openpyxl-users/N9QpvfzkJIM
        if top_left_value:  # Excel tables must have all header cells filled
            df_rows[0][0] = top_left_value
    return df_rows


def add_common_styles(workbook):
    workbook.add_named_style(NamedStyle('bold', font=Font(bold=True)))


# see Table Design tab in Excel
TABLE_STYLE_BASIC = 'TableStyleLight1'
TABLE_STYLE_BASIC_ALL_BORDER = 'TableStyleLight15'
TABLE_STYLE_MEDIUM_BLACK = 'TableStyleMedium15'
TABLE_STYLE_MEDIUM_BLUE = 'TableStyleMedium16'
TABLE_STYLE_MEDIUM_RED = 'TableStyleMedium17'
TABLE_STYLE_MEDIUM_GREEN = 'TableStyleMedium18'
TABLE_STYLE_MEDIUM_PURPLE = 'TableStyleMedium19'
TABLE_STYLE_MEDIUM_AQUA = 'TableStyleMedium20'
TABLE_STYLE_MEDIUM_ORANGE = 'TableStyleMedium21'


def mark_table(sheet, ref, has_row_header=True, has_column_header=True, table_style_name=TABLE_STYLE_BASIC):
    table_style = None
    if table_style_name is not None:
        table_style = TableStyleInfo(name=table_style_name, showFirstColumn=has_column_header,
                                     showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    table = Table(displayName='Table' + uuid.uuid4().hex,  # not sure why it's called display, it's not displayed, and it must be unique https://openpyxl.readthedocs.io/en/stable/worksheet_tables.html
                  ref=ref, tableStyleInfo=table_style, headerRowCount=1 if has_row_header else 0)
    sheet.add_table(table)

    if table_style:
        # highlight headers for non-Microsoft viewers
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        if has_row_header:
            for r in sheet.iter_rows(min_row, min_row, min_col, max_col):
                for cell in r:
                    cell.style = 'bold'
        if has_column_header:
            for r in sheet.iter_rows(min_row, max_row, min_col, min_col):
                for cell in r:
                    cell.style = 'bold'


def add_image(sheet, ref, path):
    img = Image(path)
    img.anchor = ref
    sheet.add_image(img)


def get_cells(sheet, range_ref):
    for row in sheet[range_ref]:
        for cell in row:
            yield cell


def set_number_format(sheet, range_ref, num_format='#,##0.0000'):
    for cell in get_cells(sheet, range_ref):
        cell.number_format = num_format
