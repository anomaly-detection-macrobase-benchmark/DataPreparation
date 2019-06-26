import argparse
import itertools
import os
import tempfile
from operator import itemgetter

from openpyxl import Workbook
from sklearn.metrics import average_precision_score

from utils.argparse import ArgParser
from utils.collections import find_first
from utils.datasets import load_dataset, load_column
from utils.fs import load_json, list_files
import utils.xlsx as xlsx
from utils.xlsx import autofit, append_blank_rows, pandas_dataframe_to_rows, mark_table, xlref_range_count, xlref, \
    add_image
import utils.project as project

arg_parser = ArgParser(
    description='''Generates a report.''',
    formatter_class=argparse.RawDescriptionHelpFormatter,
    epilog='''Examples:
report.py ../../macrobase/alexp/output output/report.xlsx --data-dir ../../datasets''')
arg_parser.add_argument('result_dir', type=str, help='path to the directory with the benchmark result')
arg_parser.add_argument('output_file', type=str, help='path to the output file (Excel .xlsx)')
arg_parser.add_argument('--data-dir', type=str, help='path to the directory with the datasets')
arg_parser.add_argument('--sizechanges', action='store_true', help='include sizechanges.py script output')
args = arg_parser.parse_args()

results_file_paths = [f for f in list_files(args.result_dir) if f.endswith('.json')]

execution_results = [{**load_json(f),
                      **{'_file': f, '_dir': os.path.dirname(f)}}
                     for f in results_file_paths]

dataset_ids = {r['config']['dataset']['id'] for r in execution_results}
datasets = {dataset_id: load_dataset(dataset_id, args.data_dir, 'is_anomaly') for dataset_id in dataset_ids}

classifiers = [{
    'classifier': r['config']['classifier']['id'],
    'initial_parameters': r['config']['classifier']['parameters'],
    'final_parameters': r['result']['finalAlgorithmConfig']['parameters'],
    'output_file': r['result']['algorithmOutputFilePath'],
    'scores': load_column(os.path.join(r['_dir'], r['result']['algorithmOutputFilePath']), '_OUTLIER'),
    'dataset': r['config']['dataset'],
    'result': r['result'],
} for r in execution_results]
classifiers.sort(key=itemgetter('classifier'))
classifiers = {key: sorted(list(g), key=lambda x: x['output_file'])
               for key, g in itertools.groupby(classifiers, lambda item: item['classifier'])}


def compact_classifier_parameters(items):
    result = []
    for item in items:
        result_item = find_first(result, lambda x: x['final_parameters'] == item['final_parameters'])
        if result_item is None:
            result_item = item.copy()
            del result_item['output_file']
            result_item['output_files'] = []
            result.append(result_item)
        result_item['output_files'].append(item['output_file'])
    return result


workbook = Workbook()
workbook.remove(workbook.active)  # remove default sheet

xlsx.add_common_styles(workbook)


def write_datasets_sheet():
    sheet = workbook.create_sheet('Datasets')

    # main stats table
    header = ['Dataset', 'Samples', 'Dims', 'Anomalies %']
    sheet.append(header)
    for dataset_id, dataset in datasets.items():
        stats = dataset.stats
        sheet.append([dataset_id, stats.row_count, stats.column_count, stats.anomaly_count / stats.row_count * 100])
    mark_table(sheet, xlref_range_count(1, 1, len(dataset_ids) + 1, len(header)))

    # column stats for each dataset
    for dataset_id, dataset in datasets.items():
        append_blank_rows(sheet, 2)

        sheet.append([dataset_id])
        sheet.cell(sheet.max_row, 1).style = 'bold'
        table_start_row = sheet.max_row + 1
        df_rows = pandas_dataframe_to_rows(dataset.stats.columns, top_left_value='Column stats')
        for r in df_rows:
            sheet.append(r)
        mark_table(sheet, xlref_range_count(table_start_row, 1, len(df_rows), len(df_rows[0])))

    autofit(sheet)


def write_classifiers_sheet():
    sheet = workbook.create_sheet('Classifiers')

    column_index = 1
    for name, item in classifiers.items():
        sheet.cell(1, column_index, name).style = 'bold'
        compact_classifier = compact_classifier_parameters(item)
        row_index = 2
        for compact_item in compact_classifier:
            for file in compact_item['output_files']:
                sheet.cell(row_index, column_index, file).style = 'bold'
                sheet.merge_cells(start_row=row_index, end_row=row_index,
                                  start_column=column_index, end_column=column_index + 1)
                row_index += 1
            table = [['Parameter', 'Value']] + [[key, val] for key, val in compact_item['final_parameters'].items()]
            table_start_row = row_index
            for row in table:
                for i, cell in enumerate(row):
                    sheet.cell(row_index, column_index + i, cell)
                row_index += 1
            mark_table(sheet, xlref_range_count(table_start_row, column_index, len(table), len(table[0])),
                       table_style_name=xlsx.TABLE_STYLE_BASIC_ALL_BORDER, has_column_header=False)
            row_index += 2
        column_index += 3

    autofit(sheet)


def write_evaluation_sheet():
    sheet = workbook.create_sheet('Evaluation')

    sheet.cell(1, 1, 'PR AUC').style = 'bold'
    files = [it['output_file'] for it in classifiers[list(classifiers.keys())[0]]]
    pr_table_rows = [['Algorithm', *files]] + [[classifier] + len(files) * [None] for classifier in classifiers]
    for row in pr_table_rows[1:]:
        for i, file in enumerate(files):
            classifier_results = classifiers[row[0]]
            classifier = find_first(classifier_results, lambda it: it['output_file'] == file)
            if classifier is not None:
                scores = classifier['scores']
                dataset = datasets[classifier['dataset']['id']]
                labels = dataset.labels
                pr_auc = average_precision_score(labels, scores)
                row[1 + i] = pr_auc
    for row in pr_table_rows:
        sheet.append(row)
    mark_table(sheet, xlref_range_count(2, 1, len(pr_table_rows), len(pr_table_rows[0])))

    row_index = len(pr_table_rows) + 4

    project.run_script('evaluation/quality.py', [args.result_dir,
                                                 '--data-dir', args.data_dir,
                                                 '--output-dir', tmp_dir_path,
                                                 '--plot-name', '{name}',
                                                 '--title', 'PR AUC on all datasets',
                                                 '--silent'])
    add_image(sheet, xlref(row_index, 1), os.path.join(tmp_dir_path, 'auc.png'))
    row_index += 26

    if args.sizechanges:
        project.run_script('evaluation/sizechanges.py', [args.result_dir,
                                                         '--data-dir', args.data_dir,
                                                         '--output-dir', tmp_dir_path,
                                                         '--plot-name', 'size-{name}',
                                                         '--scale', 'log',
                                                         '--title', '{name} when changing dataset size',
                                                         '--silent'])
        add_image(sheet, xlref(row_index, 1), os.path.join(tmp_dir_path, 'size-time.png'))
        row_index += 25
        add_image(sheet, xlref(row_index, 1), os.path.join(tmp_dir_path, 'size-memory.png'))
        row_index += 25
        add_image(sheet, xlref(row_index, 1), os.path.join(tmp_dir_path, 'size-auc.png'))
        row_index += 26

    autofit(sheet)


with tempfile.TemporaryDirectory() as tmp_dir_path:
    write_datasets_sheet()
    write_classifiers_sheet()
    write_evaluation_sheet()

    workbook.save(args.output_file)
